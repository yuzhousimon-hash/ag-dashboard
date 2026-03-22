"""
更新 各省准入状况及工作计划_0315.xlsx
从 province_master.json 读取最新数据，写入Excel并新增"本周变化"列
"""
import openpyxl
import json
from pathlib import Path
from datetime import datetime
from copy import copy

BASE = Path(r"g:\我的云端硬盘\AG_Workspace\政务管理")
EXCEL_PATH = BASE / "各省准入状况及工作计划_0315.xlsx"
MASTER_PATH = BASE / "data" / "province_master.json"

# Load master data
with open(MASTER_PATH, "r", encoding="utf-8") as f:
    master = json.load(f)

# Build province lookup from master
prov_data = {}
for p in master:
    name = p.get("province", "")
    prov_data[name] = p

# Load Excel
wb = openpyxl.load_workbook(EXCEL_PATH)

# ====== Sheet1: 各省招标平台状态一览图 ======
ws1 = wb["各省招标平台状态一览图"]

# First, map province rows
prov_rows = {}
for r in range(1, ws1.max_row + 1):
    v = ws1.cell(r, 1).value
    if v and "省" in str(v) or (v and any(x in str(v) for x in ["北京", "上海", "天津", "重庆", "广西", "内蒙古", "西藏", "宁夏", "新疆"])):
        clean = str(v).replace("省", "").replace("市", "").replace("自治区", "").replace("壮族", "").replace("回族", "").replace("维吾尔", "").replace("特别行政区", "")
        prov_rows[clean] = r
        prov_rows[str(v)] = r

print("Found province rows:", list(prov_rows.keys())[:10], "...")

# Add "本周变化(3/20)" column - find first empty header
change_col = ws1.max_column + 1
ws1.cell(1, change_col).value = "本周变化(3/20)"
ws1.cell(1, change_col).font = openpyxl.styles.Font(bold=True, color="FF0000")

# Define this week's intelligence updates
WEEKLY_CHANGES = {
    "江苏": {
        "changes": "✅ 金针24.1元中标;\n✅ GT40挂网成功;\n✅ GT60挂网成功",
        "col_updates": {
            7: 24.1,        # 金针挂网价格
        }
    },
    "重庆": {
        "changes": "✅ GT20切换完成;\n✅ GT40切换完成;\n✅ 金滴切换完成;\n✅ 威利坦切换完成",
        "col_updates": {
            14: "已完成✅",   # 20T平台切换
            15: "已完成✅",   # 40T平台切换
        }
    },
    "内蒙古": {
        "changes": "🔴 升级为红标(价格风险)",
        "col_updates": {
            4: "红标",       # 针剂黄标→红标
        }
    },
    "宁夏": {
        "changes": "⚠️ 药品价格自查通知;\n价格治理风险预警",
        "col_updates": {}
    },
    "河南": {
        "changes": "⚠️ 黄标预警",
        "col_updates": {
            4: "黄标",
        }
    },
    "山西": {
        "changes": "✅ GT20切换完成;\n✅ GT40切换完成;\n✅ 金针切换完成;\n✅ 金滴切换完成;\n✅ 威利坦切换完成",
        "col_updates": {
            14: "已完成✅",
            15: "已完成✅",
        }
    },
    "辽宁": {
        "changes": "⚠️ 60T挂网价108.25,需关注黄标替代方案",
        "col_updates": {}
    },
    "广东": {
        "changes": "✅ 科园已攻克;\n维护期",
        "col_updates": {
            13: "已攻克",
        }
    },
    "湖北": {
        "changes": "✅ GT20/GT40切换完成;\n金针无网待推进",
        "col_updates": {
            14: "已完成✅",
            15: "已完成✅",
        }
    },
    "甘肃": {
        "changes": "✅ 全部产品切换完成",
        "col_updates": {
            14: "已完成✅",
            15: "已完成✅",
        }
    },
    "贵州": {
        "changes": "✅ 全部产品切换完成",
        "col_updates": {
            14: "已完成✅",
            15: "已完成✅",
        }
    },
    "陕西": {
        "changes": "✅ 全部产品切换完成",
        "col_updates": {
            14: "已完成✅",
            15: "已完成✅",
        }
    },
}

# Apply updates to Sheet1
updated_count = 0
yellow_fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for prov_name, update_info in WEEKLY_CHANGES.items():
    # Find matching row
    row = None
    for key, r in prov_rows.items():
        if prov_name in key or key in prov_name:
            row = r
            break
    
    if row is None:
        # Try harder - scan all rows
        for r in range(2, ws1.max_row + 1):
            v = ws1.cell(r, 1).value
            if v and prov_name in str(v):
                row = r
                break
    
    if row:
        # Write change summary
        ws1.cell(row, change_col).value = update_info["changes"]
        ws1.cell(row, change_col).fill = yellow_fill
        
        # Apply column updates
        for col, val in update_info.get("col_updates", {}).items():
            ws1.cell(row, col).value = val
            ws1.cell(row, col).fill = yellow_fill
        
        updated_count += 1
        print(f"  Updated: {prov_name} (row {row})")
    else:
        print(f"  ⚠️ Not found: {prov_name}")

# ====== Sheet2: 省份行动计划 ======
ws2 = wb["省份行动计划"]

# Add "本周变化(3/20)" column
change_col2 = ws2.max_column + 1
ws2.cell(1, change_col2).value = "本周变化(3/20)"
ws2.cell(1, change_col2).font = openpyxl.styles.Font(bold=True, color="FF0000")

# Map provinces in Sheet2 and update
for r in range(2, ws2.max_row + 1):
    prov = ws2.cell(r, 3).value  # Column C = 省份
    if prov and str(prov) in WEEKLY_CHANGES:
        change_info = WEEKLY_CHANGES[str(prov)]
        ws2.cell(r, change_col2).value = change_info["changes"]
        ws2.cell(r, change_col2).fill = yellow_fill
        
        # Update status column (col 10) for completed items
        if "全部" in change_info["changes"] or change_info["changes"].count("✅") >= 3:
            ws2.cell(r, 10).value = "✅ 本周完成"
            ws2.cell(r, 10).fill = yellow_fill
        elif "✅" in change_info["changes"]:
            ws2.cell(r, 10).value = "进行中(有进展)"
        
        # Update 健康指数 from master
        if str(prov) in prov_data:
            new_idx = prov_data[str(prov)].get("health_index")
            if new_idx:
                ws2.cell(r, 5).value = new_idx
        
        print(f"  Sheet2 updated: {prov} (row {r})")

# Save
wb.save(EXCEL_PATH)
print(f"\n✅ Excel更新完成! 共更新 {updated_count} 个省份")
print(f"   文件: {EXCEL_PATH}")
